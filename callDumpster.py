#!/usr/bin/env python3
################# Описание ##############################################################
# Этот скрипт получает информацию о доменах из DNSDumpster и сохраняет её в Excel-файл. #
# Перед использованием убедитесь, что у вас есть API-ключ от DNSDumpster и установлены  #
# необходимые библиотеки: requests и openpyxl. ########################################## 
################# Made In cred0core #####################################################
#########################################################################################
#########################################################################################

import argparse
import os
import sys
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

### Если работаете в корп среде, убедитесь что у вас есть доступ к DNSDumpster через прокси или напрямую ###
URL = "https://api.dnsdumpster.com/domain/{}"

def parse_arguments():
    parser = argparse.ArgumentParser(
        description="CLI-утилита для сбора информации о доменах из DNSDumpster и экспорта в Excel.",
        epilog="Made In cred0core"
    )
    
    # Перенос логики выбора доменов в аргументы командной строки
    domain_group = parser.add_mutually_exclusive_group(required=True)
    domain_group.add_argument(
        "-d", "--domains", 
        nargs="+", 
        help="Один или несколько доменов через пробел"
    )
    domain_group.add_argument(
        "-f", "--file", 
        help="Путь к текстовому файлу со списком доменов (по одному на строку)"
    )
    
    parser.add_argument(
        "-k", "--api-key", 
        default=os.environ.get("DNSDUMPSTER_API_KEY"),
        help="API-ключ DNSDumpster (можно задать через переменную окружения DNSDUMPSTER_API_KEY)"
    )
    parser.add_argument(
        "-o", "--output", 
        default="dnsdumpster_report.xlsx", 
        help="Имя выходного Excel-файла (по умолчанию: dnsdumpster_report.xlsx)"
    )
    parser.add_argument(
        "-p", "--proxy", 
        help="Прокси в формате http://username:password@proxy_address:port"
    )
    
    return parser.parse_args()

def main():
    args = parse_arguments()
    
    API_KEY = args.api_key
    if not API_KEY:
        print("[-] Ошибка: Не указан API-ключ. Используйте ключ -k или переменную DNSDUMPSTER_API_KEY", file=sys.stderr)
        sys.exit(1)
        
    # Сбор доменов из аргументов или файла
    if args.domains:
        DOMAINS = args.domains
    elif args.file:
        if not os.path.exists(args.file):
            print(f"[-] Ошибка: Файл {args.file} не найден.", file=sys.stderr)
            sys.exit(1)
        with open(args.file, "r", encoding="utf-8") as f:
            DOMAINS = [line.strip() for line in f if line.strip() and not line.startswith("#")]

    HEADERS = {
        "X-API-Key": API_KEY    
    }

    # Логика прокси теперь гибко управляется через аргумент -p
    Proxy = None
    if args.proxy:
        Proxy = {
            "http": args.proxy,
            "https": args.proxy
        }

    wb = Workbook()

    ws_hosts = wb.active
    ws_hosts.title = "Hosts"

    ws_txt = wb.create_sheet("TXT")
    ws_summary = wb.create_sheet("Summary")

    # -----------------------------
    # Стили
    # -----------------------------

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    thin = Side(style="thin")
    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # -----------------------------
    # Хосты 
    # -----------------------------

    headers = [
        "Domain",
        "Record Type",
        "FQDN",
        "IP",
        "Country",
        "ASN",
        "ASN Name",
        "ASN Range",
        "PTR",
        "HTTP Server",
        "HTTPS Server",
        "HTTP Title",
        "HTTPS Title",
        "Applications"
    ]

    for col, h in enumerate(headers, start=1):
        c = ws_hosts.cell(row=1, column=col)
        c.value = h
        c.fill = header_fill
        c.font = header_font
        c.border = border

    ws_txt.append(["Domain", "TXT Record"])

    for cell in ws_txt[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    ws_summary.append([
        "Domain",
        "A",
        "NS",
        "MX",
        "CNAME",
        "TXT"
    ])

    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    # -----------------------------
    # Обработка доменов 
    # -----------------------------

    for domain in DOMAINS:

        print(f"[+] {domain}")

        try:
            r = requests.get(
                URL.format(domain),
                headers=HEADERS,
                timeout=30,
                proxies=Proxy
            )
            r.raise_for_status()
            data = r.json()
        except requests.exceptions.RequestException as e:
            print(f"[-] Ошибка выполнения запроса для {domain}: {e}", file=sys.stderr)
            continue

        ##### Краткое описание #####
        ws_summary.append([
            domain,
            len(data.get("a", [])),
            len(data.get("ns", [])),
            len(data.get("mx", [])),
            len(data.get("cname", [])),
            len(data.get("txt", []))
        ])

        # TXT
        for txt in data.get("txt", []):
            ws_txt.append([domain, txt])

        # A / NS / MX / CNAME
        for rec_type in ["a", "ns", "mx", "cname"]:

            for host in data.get(rec_type, []):

                fqdn = host.get("host")

                ips = host.get("ips", [])

                if not ips:
                    ws_hosts.append([
                        domain,
                        rec_type.upper(),
                        fqdn,
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        ""
                    ])
                    continue

                for ip in ips:

                    http = ip.get("banners", {}).get("http", {})
                    https = ip.get("banners", {}).get("https", {})

                    apps = list(
                        set(
                            http.get("apps", []) +
                            https.get("apps", [])
                        )
                    )
                    ##### Сбор полей для записи в Жксельку #####
                    ws_hosts.append([
                        domain,
                        rec_type.upper(),
                        fqdn,
                        ip.get("ip"),
                        ip.get("country"),
                        ip.get("asn"),
                        ip.get("asn_name"),
                        ip.get("asn_range"),
                        ip.get("ptr"),
                        http.get("server"),
                        https.get("server"),
                        http.get("title"),
                        https.get("title"),
                        ", ".join(apps)
                    ])

    # -----------------------------
    # Работа с экселькой
    # -----------------------------

    for ws in wb.worksheets:

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for row in ws.iter_rows():

            for cell in row:
                cell.border = border

        for col in ws.columns:

            length = 0

            column = col[0].column

            for cell in col:

                try:
                    if len(str(cell.value)) > length:
                        length = len(str(cell.value))
                except:
                    pass

            ws.column_dimensions[get_column_letter(column)].width = min(length + 3, 60)

    # -----------------------------
    # Сохранение в файл dnsdumpster_report.xlsx
    # -----------------------------

    filename = args.output

    try:
        wb.save(filename)
        print(f"\nОтчет сохранен в {filename}")
    except Exception as e:
        print(f"[-] Ошибка при сохранении файла: {e}", file=sys.stderr)

if __name__ == "__main__":
    main()